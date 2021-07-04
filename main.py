
import nest_asyncio
import time
import openpyxl
from requests_html import HTMLSession
from requests_html import AsyncHTMLSession
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.select import Select

chromedriver = 'C:/Users/henry/chromedriver'
driver = webdriver.Chrome(chromedriver)
workbook = openpyxl.load_workbook(filename="Book1.xlsx")
sheet=workbook.active
nest_asyncio.apply()


def websearch():
    j = 1
    k = 1
    session = HTMLSession()
    URL = "https://www.hkstp.org/en/our-partner-companies/company-directory/"
    driver.get(URL)

    #r = session.get(URL)
    #r.html.render()
    select = Select(driver.find_element_by_id("select-type"))
    #print(select.options)
    #print ([o.text for o in select.options])
    select.select_by_visible_text("Companies in Science Park")

    for i in range(49):
        string = "toPage(" + str(i) + ")"
        driver.execute_script(string)
        time.sleep(1)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        name = soup.find_all(class_="col col-12 col-md-6 col-lg-3")
        for elem in name:
            sheet.cell(row=j, column=1).value = elem.find_all('a',href=True)[0]['href']
            j=j+1
            sheet.cell(row=k, column=2).value = elem.find_all(class_="txt-card-title")[0].text
            k=k+1
    workbook.save("Book1.xlsx")

if __name__ == '__main__':
    websearch()