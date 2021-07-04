import nest_asyncio
import time
import openpyxl
from requests_html import HTMLSession
from requests_html import AsyncHTMLSession
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.select import Select
workbook = openpyxl.load_workbook(filename="Book1.xlsx")
sheet=workbook.active
colA = sheet['A']
chromedriver = 'C:/Users/henry/chromedriver'
driver = webdriver.Chrome(chromedriver)

def websearch1():
    i=1
    for cell in colA:
        time.sleep(.1)
        url = "https://www.hkstp.org" + cell.value
        #print(url)
        driver.get(url)
        html = driver.page_source
        elem = BeautifulSoup(html, 'html.parser')
        add = tell = email = person = website = "NA"
        if (len(elem.find_all(class_="txt-with-ico ico-place")) != 0):
            add = elem.find_all(class_="txt-with-ico ico-place")[0].text
        if (len(elem.find_all(class_="txt-with-ico ico-tel")) != 0):
            tel = elem.find_all(class_="txt-with-ico ico-tel")[0].text
        if (len(elem.find_all(class_="txt-with-ico ico-email")) != 0):
            email = elem.find_all(class_="txt-with-ico ico-email")[0].text
        if (len(elem.find_all(class_="txt-with-ico ico-chat")) != 0):
            person = elem.find_all(class_="txt-with-ico ico-chat")[0].text
        if(len(elem.find_all(class_="txt-with-ico ico-star")) != 0):
            website = elem.find_all(class_="txt-with-ico ico-star")[0].find_all('a',href=True)[0]['href']

        sheet.cell(row=i, column=3).value = add
        sheet.cell(row=i, column=4).value = tel
        sheet.cell(row=i, column=5).value = email
        sheet.cell(row=i, column=6).value = person
        sheet.cell(row=i, column=7).value = website
        i=i+1
        #print(add,tel,email,person,website)
    workbook.save("Book1.xlsx")

if __name__ == '__main__':
    websearch1()